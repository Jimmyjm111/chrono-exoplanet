"""
==============================================================================
报表生成模块 (report_generator.py)
==============================================================================

【学习目标】
本模块帮助你学习Python文件操作和Excel处理：
1. openpyxl库的使用
2. 创建和写入Excel文件
3. 设置单元格样式（字体、颜色、边框）
4. 多工作表操作
5. 文件路径处理

【什么是openpyxl？】
openpyxl是Python处理Excel文件(.xlsx)的库：
- 创建新的Excel文件
- 读取和修改现有文件
- 设置单元格格式和样式
- 支持公式和图表

==============================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side,
    Alignment, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
import os

# 导入项目模块
from _01_config import REPORTS_DIR, REPORT_FILENAME
from _02_utils import ensure_directory_exists


# ==============================================================================
# 第一部分：openpyxl基础
# ==============================================================================

def demonstrate_openpyxl_basics():
    """
    演示openpyxl的基本用法
    
    【创建Excel文件的基本步骤】
    1. 创建Workbook对象
    2. 获取或创建工作表
    3. 写入数据
    4. 设置样式（可选）
    5. 保存文件
    """
    print("【openpyxl基础演示】")
    
    # 创建工作簿
    wb = Workbook()
    
    # 获取活动工作表（默认创建的第一个）
    ws = wb.active
    
    # 设置工作表名称
    ws.title = "示例数据"
    
    # 写入数据
    # 方式1：直接给单元格赋值
    ws['A1'] = "姓名"
    ws['B1'] = "年龄"
    ws['C1'] = "城市"
    
    # 方式2：使用行列索引（1开始）
    ws.cell(row=2, column=1, value="张三")
    ws.cell(row=2, column=2, value=25)
    ws.cell(row=2, column=3, value="北京")
    
    # 方式3：追加一行数据
    ws.append(["李四", 30, "上海"])
    ws.append(["王五", 28, "广州"])
    
    # 创建新的工作表
    ws2 = wb.create_sheet("第二个表")
    ws2['A1'] = "这是第二个工作表"
    
    # 不保存，只是演示
    print("  基础操作演示完成")
    
    return wb


# ==============================================================================
# 第二部分：样式设置
# ==============================================================================

def create_header_style() -> NamedStyle:
    """
    创建表头样式
    
    【NamedStyle】
    可以定义一个命名样式，然后反复使用
    包含：字体、填充、边框、对齐等
    """
    header_style = NamedStyle(name="header")
    
    # 字体设置
    # Font参数：name(字体名), size(大小), bold(加粗), color(颜色)
    header_style.font = Font(
        name='微软雅黑',
        size=11,
        bold=True,
        color='FFFFFF'  # 白色
    )
    
    # 填充设置（背景色）
    # PatternFill参数：fill_type(填充类型), fgColor(前景色)
    header_style.fill = PatternFill(
        fill_type='solid',
        fgColor='2E86AB'  # 蓝色
    )
    
    # 对齐设置
    header_style.alignment = Alignment(
        horizontal='center',  # 水平居中
        vertical='center',    # 垂直居中
    )
    
    # 边框设置
    thin_border = Side(style='thin', color='000000')
    header_style.border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    return header_style


def create_data_style() -> dict:
    """
    创建数据单元格样式
    
    返回样式组件的字典
    """
    thin_border = Side(style='thin', color='CCCCCC')
    
    return {
        'font': Font(name='微软雅黑', size=10),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        ),
    }


def create_highlight_style(is_positive: bool = True) -> PatternFill:
    """
    创建高亮样式
    
    【参数】
    is_positive: True为正面（绿色），False为负面（红色）
    """
    if is_positive:
        return PatternFill(fill_type='solid', fgColor='C6EFCE')  # 浅绿色
    else:
        return PatternFill(fill_type='solid', fgColor='FFC7CE')  # 浅红色


# ==============================================================================
# 第三部分：写入DataFrame到工作表
# ==============================================================================

def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    include_header: bool = True,
    apply_style: bool = True,
) -> None:
    """
    将DataFrame写入工作表
    
    【参数】
    ws: 工作表对象
    df: 要写入的DataFrame
    start_row: 起始行（1开始）
    start_col: 起始列（1开始）
    include_header: 是否包含表头
    apply_style: 是否应用样式
    """
    # 使用openpyxl的dataframe_to_rows函数
    # 将DataFrame转换为可迭代的行
    rows = dataframe_to_rows(df, index=False, header=include_header)
    
    # 获取样式
    data_style = create_data_style()
    
    for row_idx, row_data in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            if apply_style:
                # 第一行（表头）用特殊样式
                if row_idx == start_row and include_header:
                    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(fill_type='solid', fgColor='2E86AB')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = data_style['font']
                    cell.alignment = data_style['alignment']
                    cell.border = data_style['border']
    
    # 调整列宽
    if apply_style:
        for col_idx, column in enumerate(df.columns, start=start_col):
            col_letter = get_column_letter(col_idx)
            # 根据列名长度设置宽度（简单估算）
            max_length = max(len(str(column)), 10)
            ws.column_dimensions[col_letter].width = max_length + 2


# ==============================================================================
# 第四部分：创建费率分析报表
# ==============================================================================

def create_fee_analysis_report(
    analysis_data: Dict[str, Any],
    save_path: str = None,
) -> str:
    """
    创建完整的费率分析Excel报表
    
    【参数】
    analysis_data: 分析数据字典，包含：
        - summary: 汇总统计
        - channel_analysis: 通道分析
        - currency_analysis: 货币分析
        - raw_data: 原始交易数据（可选）
    
    【返回】
    保存的文件路径
    """
    print("\n开始生成Excel报表...")
    
    # 创建工作簿
    wb = Workbook()
    
    # 移除默认创建的工作表
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # =========================================================================
    # 工作表1：汇总概览
    # =========================================================================
    ws_summary = wb.create_sheet("汇总概览")
    
    # 添加标题
    ws_summary['A1'] = "跨境支付费率分析报告"
    ws_summary['A1'].font = Font(name='微软雅黑', size=18, bold=True, color='2E86AB')
    ws_summary.merge_cells('A1:D1')
    
    # 添加生成时间
    ws_summary['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A2'].font = Font(name='微软雅黑', size=10, italic=True)
    ws_summary.merge_cells('A2:D2')
    
    # 添加汇总数据
    if 'summary' in analysis_data:
        summary = analysis_data['summary']
        
        ws_summary['A4'] = "关键指标"
        ws_summary['A4'].font = Font(name='微软雅黑', size=14, bold=True)
        
        # 写入指标
        metrics = [
            ("总交易笔数", summary.get('total_transactions', 0), "笔"),
            ("总交易金额", summary.get('total_amount', 0), ""),
            ("总人民币金额", summary.get('total_amount_cny', 0), "CNY"),
            ("总手续费", summary.get('total_fee', 0), ""),
            ("平均交易金额", summary.get('avg_amount', 0), ""),
            ("平均手续费", summary.get('avg_fee', 0), ""),
            ("交易成功率", summary.get('success_rate', 0), "%"),
        ]
        
        row = 5
        for name, value, unit in metrics:
            ws_summary.cell(row=row, column=1, value=name)
            
            # 格式化数值
            if unit == "%":
                formatted = f"{value * 100:.2f}%"
            elif isinstance(value, float):
                formatted = f"{value:,.2f}"
            else:
                formatted = str(value)
            
            ws_summary.cell(row=row, column=2, value=formatted)
            ws_summary.cell(row=row, column=3, value=unit)
            row += 1
    
    # =========================================================================
    # 工作表2：通道分析
    # =========================================================================
    if 'channel_analysis' in analysis_data:
        ws_channel = wb.create_sheet("通道分析")
        
        # 添加标题
        ws_channel['A1'] = "各支付通道费率分析"
        ws_channel['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        ws_channel.merge_cells('A1:G1')
        
        # 写入数据
        channel_df = analysis_data['channel_analysis']
        
        # 选择要展示的列
        display_cols = [
            'channel_name', 'transaction_count', 'total_amount',
            'avg_amount', 'total_fee', 'avg_fee_rate'
        ]
        
        # 过滤存在的列
        available_cols = [col for col in display_cols if col in channel_df.columns]
        display_df = channel_df[available_cols].copy()
        
        # 重命名列
        column_names = {
            'channel_name': '通道名称',
            'transaction_count': '交易笔数',
            'total_amount': '总金额',
            'avg_amount': '平均金额',
            'total_fee': '总费用',
            'avg_fee_rate': '平均费率',
        }
        display_df = display_df.rename(columns=column_names)
        
        write_dataframe_to_sheet(ws_channel, display_df, start_row=3)
    
    # =========================================================================
    # 工作表3：货币分析
    # =========================================================================
    if 'currency_analysis' in analysis_data:
        ws_currency = wb.create_sheet("货币分析")
        
        ws_currency['A1'] = "各货币交易分析"
        ws_currency['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        ws_currency.merge_cells('A1:F1')
        
        currency_df = analysis_data['currency_analysis']
        
        display_cols = [
            'currency', 'transaction_count', 'total_amount',
            'avg_amount', 'exchange_rate', 'total_amount_cny'
        ]
        available_cols = [col for col in display_cols if col in currency_df.columns]
        display_df = currency_df[available_cols].copy()
        
        column_names = {
            'currency': '货币',
            'transaction_count': '交易笔数',
            'total_amount': '总金额',
            'avg_amount': '平均金额',
            'exchange_rate': '汇率',
            'total_amount_cny': '人民币金额',
        }
        display_df = display_df.rename(columns=column_names)
        
        write_dataframe_to_sheet(ws_currency, display_df, start_row=3)
    
    # =========================================================================
    # 工作表4：交易明细（可选，如果数据量不大）
    # =========================================================================
    if 'raw_data' in analysis_data:
        raw_df = analysis_data['raw_data']
        
        # 限制行数，避免文件过大
        max_rows = 1000
        if len(raw_df) > max_rows:
            print(f"  交易数据超过{max_rows}条，只写入前{max_rows}条")
            raw_df = raw_df.head(max_rows)
        
        ws_detail = wb.create_sheet("交易明细")
        
        ws_detail['A1'] = "交易明细数据"
        ws_detail['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        
        # 选择要展示的列
        detail_cols = [
            'transaction_id', 'currency', 'channel_id', 'amount',
            'fee', 'fee_rate', 'status', 'created_at'
        ]
        available_cols = [col for col in detail_cols if col in raw_df.columns]
        
        write_dataframe_to_sheet(
            ws_detail,
            raw_df[available_cols],
            start_row=3
        )
    
    # =========================================================================
    # 保存文件
    # =========================================================================
    if save_path is None:
        ensure_directory_exists(REPORTS_DIR)
        save_path = os.path.join(REPORTS_DIR, REPORT_FILENAME)
    
    wb.save(save_path)
    print(f"  Excel报表已保存: {save_path}")
    
    return save_path


# ==============================================================================
# 第五部分：添加图表到Excel
# ==============================================================================

def add_chart_to_sheet(
    ws,
    chart_type: str,
    data_range: str,
    categories_range: str,
    title: str,
    position: str = "E2",
) -> None:
    """
    在工作表中添加图表
    
    【openpyxl图表】
    支持多种图表类型：BarChart, PieChart, LineChart等
    
    【参数】
    ws: 工作表对象
    chart_type: 图表类型 ('bar', 'pie')
    data_range: 数据范围
    categories_range: 类别范围
    title: 图表标题
    position: 图表位置（单元格地址）
    """
    if chart_type == 'bar':
        chart = BarChart()
        chart.type = "col"  # 柱状图
        chart.style = 10
    elif chart_type == 'pie':
        chart = PieChart()
    else:
        return
    
    chart.title = title
    chart.width = 15
    chart.height = 10
    
    # 添加到工作表
    ws.add_chart(chart, position)


# ==============================================================================
# 第六部分：快捷报表生成函数
# ==============================================================================

def generate_simple_report(
    df: pd.DataFrame,
    filename: str = "简易报表.xlsx",
) -> str:
    """
    生成简易报表（快速生成）
    
    适合快速导出数据到Excel
    """
    ensure_directory_exists(REPORTS_DIR)
    save_path = os.path.join(REPORTS_DIR, filename)
    
    # 使用pandas直接导出（最简单的方式）
    # 但这样没有样式
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='数据', index=False)
    
    print(f"  简易报表已保存: {save_path}")
    return save_path


def generate_styled_report(
    df: pd.DataFrame,
    title: str = "数据报表",
    filename: str = "样式报表.xlsx",
) -> str:
    """
    生成带样式的报表
    """
    ensure_directory_exists(REPORTS_DIR)
    save_path = os.path.join(REPORTS_DIR, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"
    
    # 添加标题
    ws['A1'] = title
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    
    # 添加时间
    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(name='微软雅黑', size=9, italic=True)
    
    # 写入数据
    write_dataframe_to_sheet(ws, df, start_row=4)
    
    wb.save(save_path)
    print(f"  样式报表已保存: {save_path}")
    
    return save_path


# ==============================================================================
# 学习小结
# ==============================================================================
"""
【本模块学到的内容】

1. openpyxl基础:
   wb = Workbook()                    # 创建工作簿
   ws = wb.active                     # 获取活动工作表
   ws = wb.create_sheet("名称")       # 创建新工作表
   ws['A1'] = "值"                    # 写入单元格
   ws.cell(row=1, column=1, value=x)  # 按行列写入
   ws.append([列表])                  # 追加一行
   wb.save("文件名.xlsx")             # 保存文件

2. 单元格样式:
   Font: 字体（name, size, bold, color）
   PatternFill: 填充/背景色
   Border: 边框
   Alignment: 对齐方式

3. 常用操作:
   ws.merge_cells('A1:D1')           # 合并单元格
   ws.column_dimensions['A'].width   # 设置列宽
   get_column_letter(n)              # 数字转列字母

4. DataFrame导出:
   dataframe_to_rows(df)             # 转换为行
   pd.ExcelWriter()                  # pandas方式导出

【练习建议】
1. 添加条件格式（如费率高于阈值标红）
2. 在Excel中添加公式（如求和、平均）
3. 添加数据验证（下拉列表）
"""


# ==============================================================================
# 模块测试代码
# ==============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("报表生成模块测试")
    print("=" * 60)
    
    # 演示基础
    demonstrate_openpyxl_basics()
    
    # 生成测试数据
    from _04_data_generator import generate_transactions
    from _06_analyzer import (
        transactions_to_dataframe,
        analyze_by_channel,
        analyze_by_currency,
        calculate_summary_statistics
    )
    
    print("\n生成测试数据...")
    transactions = generate_transactions(count=200)
    df = transactions_to_dataframe(transactions)
    
    # 准备分析数据
    analysis_data = {
        'summary': calculate_summary_statistics(df),
        'channel_analysis': analyze_by_channel(df),
        'currency_analysis': analyze_by_currency(df),
        'raw_data': df,
    }
    
    # 生成报表
    print("\n生成完整报表...")
    report_path = create_fee_analysis_report(analysis_data)
    
    # 生成简易报表
    print("\n生成简易报表...")
    generate_simple_report(df, "交易数据导出.xlsx")
    
    print("\n" + "=" * 60)
    print("报表生成完成！")
    print(f"报表目录: {os.path.abspath(REPORTS_DIR)}")
